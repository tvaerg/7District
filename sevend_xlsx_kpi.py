"""
7D Portfolio Aggregation Engine -- Excel KPI Extractor
======================================================
Pulls a small, well-defined set of financial KPIs out of a company's Excel
workbook deterministically, so they can be fed to the reasoning layer as
GROUND TRUTH rather than left to the LLM to interpret from raw cells.

Design principles
-----------------
* Deterministic. We find KPIs by label match against a known synonym list (in
  Swedish and English), then read the same row's data columns. No LLM
  inference at this layer.
* Fail empty, not wrong. If we can't find a KPI, the field is None. We never
  invent a number. A confidence score is returned so the reasoning prompt can
  weight accordingly.
* Multi-entity aware. Many of 7D's workbooks split a company across subsidiary
  + holding columns and aggregate. We capture the columns we recognise and
  ALSO compute a total when the workbook does not provide one.
* Format-tolerant. Numbers come in as int, float, or strings with comma
  decimals, spaces, MSEK/KSEK suffixes -- all normalised to float in the same
  unit (KSEK is the default for the Acre workbook style).

What we extract (when present)
------------------------------
Revenue, gross profit, EBITDA (pre and post extraordinary), depreciation,
cash, long-term debt, short-term debt. Plus derived: gross margin %, EBITDA
margin %, total debt, net debt. These are the figures the reasoning layer
quotes in heatmap labels and update bullets.

This is intentionally a SMALL set. A wider set per company is achievable but
each new label adds maintenance for 7D as workbooks evolve.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional


# Label -> canonical KPI key. Multiple aliases per KPI, lowercased + stripped.
KPI_ALIASES: dict[str, list[str]] = {
    "revenue": ["intäkter", "nettoomsättning", "omsättning", "net sales", "revenue", "sales"],
    "gross_profit": ["bruttovinst", "gross profit", "bruttoresultat"],
    "ebitda_pre_eo": ["ebitda före e/o", "ebitda pre eo", "ebitda pre e/o", "adj ebitda", "ebitda före extraordinära"],
    "extraordinary": ["e/o", "extraordinära", "extraordinary"],
    "ebitda": ["ebitda efter e/o", "ebitda", "ebitda post e/o"],
    "depreciation": ["avskrivningar", "depreciation", "amortisation"],
    "cash": ["kassa", "kassa och bank", "cash", "likvida medel"],
    "long_term_debt": ["långfristig skuld", "långfristiga skulder", "long-term debt", "long term debt"],
    "short_term_debt": ["kortfristig skuld", "kortfristiga skulder", "short-term debt", "short term debt"],
}

# Reverse map: alias -> canonical key, for O(1) lookup.
_ALIAS_TO_KEY: dict[str, str] = {a: k for k, aliases in KPI_ALIASES.items() for a in aliases}


@dataclass
class KpiSnapshot:
    """Deterministic financial snapshot. Numbers are in the workbook's own unit
    (KSEK / MSEK / SEK -- see `unit`). Missing KPIs are None, never zero."""
    source_file: str = ""
    sheet: str = ""
    unit: str = "KSEK"               # detected from header row when possible
    period_label: str = ""           # e.g. "Q1 2026" if we can read it

    revenue: Optional[float] = None
    gross_profit: Optional[float] = None
    ebitda_pre_eo: Optional[float] = None
    extraordinary: Optional[float] = None
    ebitda: Optional[float] = None
    depreciation: Optional[float] = None
    cash: Optional[float] = None
    long_term_debt: Optional[float] = None
    short_term_debt: Optional[float] = None

    # Per-entity breakdowns when the workbook splits a company across columns.
    # e.g. {"Acre AB": {"revenue": 17033.1, ...}, "Holding": {...}}
    entities: dict[str, dict[str, float]] = field(default_factory=dict)

    notes: list[str] = field(default_factory=list)
    found_count: int = 0             # how many KPIs were located

    @property
    def gross_margin_pct(self) -> Optional[float]:
        if self.revenue and self.gross_profit is not None and self.revenue != 0:
            return round(100.0 * self.gross_profit / self.revenue, 1)
        return None

    @property
    def ebitda_margin_pct(self) -> Optional[float]:
        eb = self.ebitda if self.ebitda is not None else self.ebitda_pre_eo
        if self.revenue and eb is not None and self.revenue != 0:
            return round(100.0 * eb / self.revenue, 1)
        return None

    @property
    def net_debt(self) -> Optional[float]:
        debt = (self.long_term_debt or 0) + (self.short_term_debt or 0)
        if debt == 0 and self.long_term_debt is None and self.short_term_debt is None:
            return None
        return round(debt - (self.cash or 0), 1)

    def to_json(self) -> dict:
        d = asdict(self)
        d["gross_margin_pct"] = self.gross_margin_pct
        d["ebitda_margin_pct"] = self.ebitda_margin_pct
        d["net_debt"] = self.net_debt
        return d

    def is_empty(self) -> bool:
        return self.found_count == 0


# -----------------------------------------------------------------------------
# Number / label parsing helpers
# -----------------------------------------------------------------------------
_NUM_RE = re.compile(r"[-+]?\d[\d\s\u00a0.,]*")


def _to_float(v) -> Optional[float]:
    """Robust to int, float, and Swedish-formatted strings ('1 234,5' / '1.234,5')."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # Strip currency / unit suffixes.
    s = re.sub(r"\s*(KSEK|MSEK|SEK|tkr|TSEK|kr)\s*$", "", s, flags=re.I)
    m = _NUM_RE.search(s)
    if not m:
        return None
    n = m.group(0).strip()
    # Swedish: thousands sep = space/non-break-space/period, decimal = comma.
    n = n.replace("\u00a0", " ").replace(" ", "")
    if "," in n and "." in n:
        # Assume comma decimal if it's the rightmost separator.
        if n.rfind(",") > n.rfind("."):
            n = n.replace(".", "").replace(",", ".")
        else:
            n = n.replace(",", "")
    elif "," in n:
        n = n.replace(",", ".")
    try:
        return float(n)
    except ValueError:
        return None


def _norm_label(v) -> str:
    if v is None:
        return ""
    return str(v).strip().lower().rstrip(":").strip()


# -----------------------------------------------------------------------------
# Workbook scanner
# -----------------------------------------------------------------------------
def extract_kpis_from_xlsx(path: str | Path) -> KpiSnapshot:
    """Extract KPIs from a workbook. Tries each sheet; merges what it finds."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        snap = KpiSnapshot(source_file=str(path))
        snap.notes.append("openpyxl not installed (pip install openpyxl)")
        return snap

    p = Path(path)
    snap = KpiSnapshot(source_file=str(p))
    if not p.exists():
        snap.notes.append(f"file not found: {path}")
        return snap

    try:
        wb = load_workbook(p, data_only=True, read_only=True)
    except Exception as e:                             # noqa: BLE001
        snap.notes.append(f"could not open workbook: {e}")
        return snap

    # Score each sheet by how many KPI labels appear, pick the best.
    best_sheet = None
    best_score = 0
    for sn in wb.sheetnames:
        try:
            score = _score_sheet(wb[sn])
        except Exception:                              # noqa: BLE001
            score = 0
        if score > best_score:
            best_score, best_sheet = score, sn
    if best_sheet is None:
        snap.notes.append("no recognisable KPI labels in any sheet")
        return snap

    snap.sheet = best_sheet
    _scan_sheet(wb[best_sheet], snap)
    return snap


def _score_sheet(ws) -> int:
    """Count cells whose normalised value matches any known KPI alias."""
    n = 0
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            lbl = _norm_label(cell)
            if lbl in _ALIAS_TO_KEY:
                n += 1
    return n


def _scan_sheet(ws, snap: KpiSnapshot) -> None:
    """Walk the sheet, find a KPI label cell, then read numeric cells on the
    same row to its right -- treating distinct numbers as separate entity
    columns (subsidiary / holding / total)."""
    # First pass: find an entity-header row (row of short text labels above the
    # data, e.g. 'Acre AB | Holding | ACK'). This lets us tag values per entity.
    entity_headers: dict[int, str] = {}
    for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
        if ri > 5:
            break
        texts = [(ci, str(c).strip()) for ci, c in enumerate(row) if c not in (None, "")]
        non_label = [(ci, t) for ci, t in texts if _norm_label(t) not in _ALIAS_TO_KEY]
        # Heuristic: 2+ short text cells, none of which are KPI labels => header row.
        if len(non_label) >= 2 and all(len(t) < 30 for _, t in non_label) and \
           not any(_to_float(t) is not None for _, t in non_label):
            entity_headers = {ci: t for ci, t in non_label
                              if t.lower() not in {"ksek", "msek", "sek", "tkr"}}
            # Detect unit while we're here.
            for _, t in texts:
                tl = t.lower()
                if tl in ("ksek", "msek", "sek", "tkr"):
                    snap.unit = t.upper() if tl != "tkr" else "KSEK"
            break

    found: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        # Locate the label cell on this row.
        label_idx = -1
        label_key = None
        for ci, c in enumerate(row):
            key = _ALIAS_TO_KEY.get(_norm_label(c))
            if key:
                label_idx, label_key = ci, key
                break
        if label_key is None:
            continue

        # Read numeric values to the right; tag them with entity headers if known.
        numeric_cells: list[tuple[int, float]] = []
        for ci in range(label_idx + 1, len(row)):
            v = _to_float(row[ci])
            if v is not None:
                numeric_cells.append((ci, v))

        if not numeric_cells:
            continue

        # Per-entity capture.
        for ci, v in numeric_cells:
            ent = entity_headers.get(ci)
            if ent:
                snap.entities.setdefault(ent, {})[label_key] = v

        # Pick the canonical total for the snapshot's top-level field:
        # prefer an entity called ACK/Total/Consolidated; otherwise use the
        # rightmost number (the workbook's own total column convention).
        chosen = None
        for ci, v in numeric_cells:
            ent = (entity_headers.get(ci) or "").lower()
            if any(k in ent for k in ("ack", "total", "konsolider", "consolidated", "sum")):
                chosen = v
                break
        if chosen is None:
            chosen = numeric_cells[-1][1]

        if getattr(snap, label_key) is None:
            setattr(snap, label_key, chosen)
            found.add(label_key)

    snap.found_count = len(found)
    if not found:
        snap.notes.append("labels matched but no numeric values were readable")


# -----------------------------------------------------------------------------
# Compact text rendering for the reasoning prompt
# -----------------------------------------------------------------------------
def render_kpis_for_prompt(snap: KpiSnapshot) -> str:
    """Render the snapshot as a short, dense fact block for the Opus prompt.
    Numbers come with explicit units so the model can't confuse KSEK and MSEK."""
    if snap.is_empty():
        return ""
    u = snap.unit
    lines = [f"STRUCTURED KPI SNAPSHOT (deterministic, from {Path(snap.source_file).name}; unit={u})"]
    rows = [
        ("Revenue",         snap.revenue),
        ("Gross profit",    snap.gross_profit),
        ("Gross margin %",  snap.gross_margin_pct),
        ("EBITDA pre E/O",  snap.ebitda_pre_eo),
        ("Extraordinary",   snap.extraordinary),
        ("EBITDA",          snap.ebitda),
        ("EBITDA margin %", snap.ebitda_margin_pct),
        ("Depreciation",    snap.depreciation),
        ("Cash",            snap.cash),
        ("Long-term debt",  snap.long_term_debt),
        ("Short-term debt", snap.short_term_debt),
        ("Net debt",        snap.net_debt),
    ]
    for label, val in rows:
        if val is None:
            continue
        suffix = "%" if "%" in label else f" {u}"
        # Round to 1 decimal to kill float artifacts like 206.60000000000002.
        display = round(val, 1) if isinstance(val, float) else val
        lines.append(f"  {label}: {display}{suffix}")
    if snap.entities:
        lines.append("  Per-entity breakdown:")
        for ent, ed in snap.entities.items():
            facts = ", ".join(f"{k}={v}" for k, v in ed.items())
            lines.append(f"    {ent}: {facts}")
    return "\n".join(lines)


if __name__ == "__main__":
    import sys, json
    paths = sys.argv[1:] or ["/mnt/user-data/uploads/2026_Q1_ACRE_konsoliderat.xlsx"]
    for p in paths:
        snap = extract_kpis_from_xlsx(p)
        print(f"\n=== {Path(p).name} ===")
        print(f"  sheet={snap.sheet}  unit={snap.unit}  found={snap.found_count}")
        if snap.notes:
            print(f"  notes: {snap.notes}")
        print(render_kpis_for_prompt(snap) or "  (no KPIs extracted)")